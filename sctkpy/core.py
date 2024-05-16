# -*- coding: utf-8 -*-

import csv
import datetime
import os
import shutil

import openpyxl

from .config import (
    default_roster_csv_path,
    files_hit_list_template_path,
    grade_report_template_path,
    no_punting,
    no_study_hours,
    proj_root_dir,
    social_probation,
    study_checks_template_path,
    tier_one,
    tier_two,
)
from .db_funcs import get_files_due_list_db
from .member import Member


def parse_roster_report(
    csv_file_path: str | None = None, rows_to_header: int = 3
) -> tuple[str, list[tuple[str, str, str]]]:
    """Parses the "greeklife_roster_report.csv" either from a given path or
    via the root of the project directory. Saves the results and inserts/
    updates member information in SQLite.

    Args:
        csv_file_path (str | None, optional):
            Optional path to csv file. Defaults to None.
        rows_to_header (int, optional):
            Rows to skip to get the the header. Defaults to 3.

    Returns:
        Roster term,
        List of strings of First Name, Last Name, and In/Out House status
    """

    if csv_file_path is None:
        csv_file_path = default_roster_csv_path

    with open(csv_file_path, "r", encoding="utf-8") as csv_file:
        for _ in range(rows_to_header):  # Skip first rows to get to head
            next(csv_file)
        dr = csv.DictReader(csv_file)
        raw_member_info: list[tuple[str, str, str, str]] = sorted(
            list(
                {
                    (
                        row["First Name"],
                        row["Last Name"],
                        row["in/out House"],
                        row["Term"],
                    )
                    for row in dr
                }
            ),
            key=lambda tup: (1 if tup[2] == "IN" else 0, tup[1]),
        )
        member_info: list[tuple[str, str, str]] = [
            row[:3] for row in raw_member_info
        ]
        assert (
            len({row[3] for row in raw_member_info}) == 1
        ), "More than one unique term in roster"
        term = {row[3] for row in raw_member_info}.pop()

    return term, member_info


def get_member_info() -> list[Member]:
    """Gets list of dudes from grade report and spits them out

    Returns:
        list[Member]: List of Member objects
    """
    _, member_info = parse_roster_report()
    members = [
        Member(
            name=" ".join(member[:2]),
            out_of_house=True if member[2] == "OUT" else False,
        )
        for member in member_info
    ]

    return sorted(
        members,
        key=lambda dude: (
            not dude.out_of_house,
            (
                int(dude.pledge_class[2:])
                if dude.pledge_class is not None
                else float("inf")  # if no pledge class, goes last
            ),
            (
                (0 if dude.pledge_class[:2] == "SP" else 1)
                if dude.pledge_class is not None
                else float("inf")  # if no pledge class, goes last
            ),
            dude.name.split()[1],
        ),  # Sort by year, semester (SP/FS), then last name
    )


def generate_grade_report(term: str) -> None:
    """Make a copy of and write data to a grade report for grade data from a
    given term

    Args:
        term (str): Term to use grade data from
    """
    grade_report_save_path: str = os.path.join(
        proj_root_dir, f"{term}_Grade_Report.xlsx"
    )
    shutil.copy(src=grade_report_template_path, dst=grade_report_save_path)

    workbook = openpyxl.load_workbook(filename=grade_report_save_path)
    worksheet = workbook.active

    # Write member grade data
    start_row = 2
    start_col = 1
    members: list[Member] = get_member_info()

    statistics_data: dict = {"pledge_class": [], "gpa": []}

    for i, member in enumerate(members):  # Write columns
        ret_term, term_gpa, study_hours = member.study_hours(term)

        if term_gpa is not None:
            statistics_data["pledge_class"].append(member.pledge_class)
            statistics_data["gpa"].append(term_gpa)

        rounded_term_gpa = (
            float(f"{term_gpa:.3f}") if term_gpa is not None else None
        )

        for j, datum in enumerate(
            (
                member.name,
                ret_term,
                member.pledge_class,
                rounded_term_gpa,
                study_hours,
            )
        ):  # Write row
            worksheet.cell(
                row=start_row + i, column=start_col + j, value=datum
            )

    # House GPA Statistics

    # Whole house average GPA
    house_avg_gpa: float = sum(statistics_data["gpa"]) / len(
        statistics_data["gpa"]
    )
    start_row = 17
    start_col = 9
    worksheet.cell(
        row=start_row, column=start_col, value=f"{house_avg_gpa:.3f}"
    )

    unique_pledge_classes = []
    for pledge_class in statistics_data["pledge_class"]:
        if pledge_class not in unique_pledge_classes:
            unique_pledge_classes.append(pledge_class)

    # Pledge class average GPA
    start_row = 21
    start_col = 7
    for i, pledge_class in enumerate(unique_pledge_classes):
        pc_grade_data = [
            statistics_data["gpa"][i]
            for i, pc in enumerate(statistics_data["pledge_class"])
            if pc == pledge_class
        ]
        pc_avg_grade = sum(pc_grade_data) / len(pc_grade_data)

        # Write data
        worksheet.cell(row=start_row + i, column=start_col, value=pledge_class)
        worksheet.cell(
            row=start_row + i, column=start_col + 1, value=len(pc_grade_data)
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 2,
            value=f"{pc_avg_grade:.3f}",
        )

    # Add hidden numbers used for comparison in the xlsx conditional formatting
    # In column F, yellow => tier one (F4), red => tier two (F5).
    start_row = 4
    start_col = 6
    for i, option in enumerate((tier_one, tier_two)):
        worksheet.cell(
            row=start_row + i,
            column=start_col,
            value=option.bound,
        )

    # In-House - Key Info
    start_row = 4
    start_col = 7
    for i, option in enumerate(
        (no_study_hours, tier_one, tier_two, no_punting)
    ):
        worksheet.cell(
            row=start_row + i,
            column=start_col,
            value=option.desc_in_house,
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 1,
            value=option.result_in_house,
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 2,
            value=f"GPA {option.condition} {option.bound}",
        )

    # Out-House - Key Info
    start_row = 10
    start_col = 7
    for i, option in enumerate((no_study_hours, tier_one, tier_two)):
        worksheet.cell(
            row=start_row + i,
            column=start_col,
            value=option.desc_out_house,
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 1,
            value=option.result_out_house,
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 2,
            value=f"GPA {option.condition} {option.bound}",
        )

    # Social Probation - Key Info
    start_row = 15
    start_col = 7
    option = social_probation
    worksheet.cell(
        row=start_row,
        column=start_col,
        value=option.desc_in_house,
    )
    worksheet.cell(
        row=start_row,
        column=start_col + 1,
        value=option.result_in_house,
    )
    worksheet.cell(
        row=start_row,
        column=start_col + 2,
        value=f"GPA {option.condition} {option.bound}",
    )

    workbook.save(grade_report_save_path)
    os.startfile(grade_report_save_path)

    return


def generate_study_check_sheet(term: str) -> None:
    """Make a copy of and write data to a study-checklist for grade data from
    a given term

    Args:
        term (str): Term to use grade data from
    """
    study_checks_save_path: str = os.path.join(
        proj_root_dir, f"{term}_Study_Check_Sheet.xlsx"
    )
    shutil.copy(src=study_checks_template_path, dst=study_checks_save_path)

    workbook = openpyxl.load_workbook(filename=study_checks_save_path)
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value=f"{term}, Week of:")

    start_row = 3
    start_col = 1
    members: list[Member] = get_member_info()

    i = 0
    for member in members:  # Write columns
        _, _, study_hours = member.study_hours(term)
        if len(study_hours) > 0 and not member.out_of_house:
            worksheet.cell(
                row=start_row + i, column=start_col, value=member.name
            )
            worksheet.cell(
                row=start_row + i, column=start_col + 1, value=study_hours
            )
            i += 1

    workbook.save(study_checks_save_path)
    os.startfile(study_checks_save_path)

    return


def generate_files_responsibility_report(term: str) -> None:
    """Make a copy of and write data to a files turn-in checklist from course
    data from a given term

    Args:
        term (str): Term to use course data from
    """
    files_hit_list_save_path: str = os.path.join(
        proj_root_dir, f"{term}_Files_Hit_List.xlsx"
    )
    shutil.copy(src=files_hit_list_template_path, dst=files_hit_list_save_path)

    workbook = openpyxl.load_workbook(filename=files_hit_list_save_path)
    worksheet = workbook.active

    files_due_list = get_files_due_list_db(term)

    # Title information
    start_row = 1
    start_col = 1
    worksheet.cell(
        row=start_row,
        column=start_col,
        value=f"File Responsibility Report for: {term}",
    )
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    worksheet.cell(
        row=start_row + 1,
        column=start_col,
        value=f"Report Generated at {current_time}",
    )
    worksheet.cell(
        row=start_row + 2,
        column=start_col,
        value=f"Total: {len(files_due_list)} files",
    )

    # Files Due List Data
    start_row = 6
    start_col = 1
    i = 0
    for class_name, class_num, dudes in files_due_list:
        worksheet.cell(
            row=start_row + i,
            column=start_col,
            value=class_name,
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 1,
            value=class_num,
        )
        i += 1
        for dude in dudes.split(","):
            worksheet.cell(
                row=start_row + i,
                column=start_col + 4,
                value=dude,
            )
            i += 1

    workbook.save(files_hit_list_save_path)
    os.startfile(files_hit_list_save_path)

    return
