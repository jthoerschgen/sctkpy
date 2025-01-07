"""Object for generating Study Check Reports"""

import os
import shutil
from datetime import datetime

import openpyxl

from sctkpy.config import proj_root_dir, study_checks_template_path
from sctkpy.member.member import GradeReportItem
from sctkpy.reports.grade_report import GradeReport


class StudyCheckReport(GradeReport):
    """Extra logic for writing a Study Check Report from a GradeReport object.

    Attributes:
        template_path (str): Path of the template spreadsheet.
        save_dir (str): Directory path of where the report is saved.
        report_data (list[GradeReportItem]): List of GradeReportItems.
            Defaults to an empty list.
        workbook (openpyxl.Workbook): The workbook where the report is written
            to. Defaults to None.
    """

    def __init__(self, roster_report_path: str, save_dir: str = proj_root_dir):
        super().__init__(roster_report_path)
        self.template_path: str = study_checks_template_path
        self.save_dir: str = save_dir
        self.report_data: list[GradeReportItem] = []
        self.workbook: openpyxl.Workbook | None = None

    def __call__(
        self, selected_term: str, open_on_finish: bool = True
    ) -> None:
        current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        save_path: str = os.path.join(
            self.save_dir, f"Grade_Report_{selected_term}_{current_time}.xlsx"
        )

        self._copy_template(save_path)
        self._write_member_data(selected_term)
        self._save(save_path, open_on_finish)

    def _copy_template(self, save_path: str) -> None:
        """Makes a copy of the study check report template.

        Args:
            save_path (str): The path where the copy is saved.

        Returns:
            None: None
        """
        shutil.copy(src=self.template_path, dst=save_path)
        self.workbook = openpyxl.load_workbook(save_path)

    def _organize_report_data(self, selected_term: str) -> None:
        """Organizes the report information.

        Args:
            selected_term (str): The desired term to organize data from.

        Returns:
            None: None
        """

        def sort_key(
            member_data: GradeReportItem,
        ) -> tuple[bool, float | int, float | int, str]:
            """Function used for sorting grade report row values.

            Rows are sorted by: out-of-house before in-house, pledge class
            year, pledge class semester (Spring before Fall), member last name.

            Gives the row value from a list a number value used for sorting.

            Args:
                member_data (GradeReportItem): The dataclass that represents a
                    row that is being sorted.

            Returns:
                tuple[bool, float | int, float | int, str]: The sorting value
                    for the row.
            """
            pledge_class_year_order: float | int = float("inf")
            pledge_class_semester_order: float | int = float("inf")
            if member_data.pledge_class is not None:
                pledge_class_year_order = int(member_data.pledge_class[2:])
                if member_data.pledge_class[:2] == "SP":
                    pledge_class_semester_order = 0
                else:
                    pledge_class_semester_order = 1
            return (
                not member_data.out_of_house,
                pledge_class_year_order,
                pledge_class_semester_order,
                member_data.name.split()[1],
            )

        report_data: list[GradeReportItem] = []
        for member in self.members.values():
            member_data: GradeReportItem = member.determine_academic_probation(
                selected_term
            )
            report_data.append(member_data)
        self.report_data = sorted(report_data, key=sort_key)

    def _write_member_data(self, selected_term: str) -> None:
        """Writes the report data.

        Args:
            selected_term (str): The desired term to write study hour data
                from.

        Raises:
            ValueError: When the method is called when the workbook attribute
                is None.

        Returns:
            None: None
        """
        self._organize_report_data(selected_term)

        if self.workbook is None:
            raise ValueError("Workbook cannot be None, copy template first.")
        worksheet = self.workbook.active

        start_row = 3
        start_column = 1

        i: int = 0
        for member in self.report_data:
            if member.out_of_house or len(member.study_hours) == 0:
                continue
            worksheet.cell(
                row=start_row + i,
                column=start_column,
                value=member.name,
            )
            worksheet.cell(
                row=start_row + i,
                column=start_column + 1,
                value=member.study_hours,
            )
            i += 1

    def _save(self, save_path: str, open_file: bool = False) -> None:
        """Saves the report to file.

        Args:
            save_path (str): The file path where the report is saved to.
            open_file (bool, optional): Option to open the file after it is
                saved. Defaults to False.

        Raises:
            ValueError: When the method is called when the workbook attribute
                is None.

        Returns:
            None: None
        """
        if self.workbook is None:
            raise ValueError("Workbook cannot be None, copy template first.")
        self.workbook.save(save_path)
        if open_file:
            os.startfile(save_path)
