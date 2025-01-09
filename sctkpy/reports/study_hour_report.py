"""Object for generating Study Hour Reports"""

import os
import shutil
from datetime import datetime

import openpyxl

from sctkpy.config import (
    academic_suspension_cumulative,
    academic_suspension_term,
    get_off_no_punting,
    get_off_social_probation,
    grade_report_template_path,
    no_punting,
    no_study_hours,
    reduce_one_tier,
    reduce_two_tiers,
    social_probation,
    tier_one,
    tier_two,
)
from sctkpy.member.member import GradeReportItem
from sctkpy.member.term import Semester
from sctkpy.reports.grade_report import GradeReport


class StudyHourReport(GradeReport):
    """Extra logic for writing a Study Hour Report from a GradeReport object.

    Attributes:
        template_path (str): Path of the template spreadsheet.
        save_dir (str): Directory path of where the report is saved.
        report_data (list[GradeReportItem]): List of GradeReportItems.
            Defaults to an empty list.
        workbook (openpyxl.Workbook): The workbook where the report is written
            to. Defaults to None.
    """

    def __init__(self, roster_report_path: str, save_dir: str):
        super().__init__(roster_report_path)
        self.template_path: str = grade_report_template_path
        self.save_dir: str = save_dir
        self.report_data: list[GradeReportItem] = []
        self.workbook: openpyxl.Workbook | None = None

    def __call__(
        self,
        selected_term: str,
        open_on_finish: bool = True,
    ) -> None:
        current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        save_path: str = os.path.join(
            self.save_dir, f"Grade_Report_{selected_term}_{current_time}.xlsx"
        )

        self._copy_template(save_path)
        self._write_key_information()
        self._write_member_data(selected_term)
        self._write_statistics(selected_term)
        self._save(save_path, open_on_finish)

    def _copy_template(self, save_path: str) -> None:
        """Makes a copy of the study hour report template.

        Args:
            save_path (str): The path where the copy is saved.

        Returns:
            None: None
        """
        shutil.copy(src=self.template_path, dst=save_path)
        self.workbook = openpyxl.load_workbook(save_path)

    def _write_key_information(self) -> None:
        """Writes information to the key of the study hour report.

        Raises:
            ValueError: When the method is called when the workbook attribute
                is None.

        Returns:
            None: None
        """
        if self.workbook is None:
            raise ValueError("Workbook cannot be None, copy template first.")
        worksheet = self.workbook.active

        # Hidden numbers used for conditional formatting in the template sheet.
        #   -> Yellow Cells, Tier One (F4)
        #   -> Red Cells, Tier Two (F5)
        start_row = 4
        start_column = 9
        for i, format_value in enumerate((tier_one, tier_two)):
            worksheet.cell(
                row=start_row + i,
                column=start_column,
                value=format_value.bound,
            )

        # In-House Key Information
        start_row = 4
        start_column = 10
        for i, in_house_tier in enumerate(
            (no_study_hours, tier_one, tier_two, no_punting)
        ):
            worksheet.cell(
                row=start_row + i,
                column=start_column,
                value=in_house_tier.description_in_house,
            )
            worksheet.cell(
                row=start_row + i,
                column=start_column + 3,
                value=in_house_tier.result_in_house,
            )
            worksheet.cell(
                row=start_row + i,
                column=start_column + 4,
                value=f"GPA {in_house_tier.condition} {in_house_tier.bound}",
            )

        # Out-of-House Key Information
        start_row = 10
        start_column = 10
        for i, out_house_tier in enumerate(
            (no_study_hours, tier_one, tier_two)
        ):
            worksheet.cell(
                row=start_row + i,
                column=start_column,
                value=out_house_tier.description_out_house,
            )
            worksheet.cell(
                row=start_row + i,
                column=start_column + 3,
                value=out_house_tier.result_out_house,
            )
            worksheet.cell(
                row=start_row + i,
                column=start_column + 4,
                value=f"GPA {out_house_tier.condition} {out_house_tier.bound}",
            )

        # Social Probation & Suspension Key Information
        start_row = 15
        start_column = 10
        for i, probation_tier in enumerate(
            (
                social_probation,
                academic_suspension_term,
                academic_suspension_cumulative,
            )
        ):
            worksheet.cell(
                row=start_row + i,
                column=start_column,
                value=probation_tier.description_in_house,
            )
            worksheet.cell(
                row=start_row + i,
                column=start_column + 3,
                value=probation_tier.result_in_house,
            )
            worksheet.cell(
                row=start_row + i,
                column=start_column + 4,
                value=f"GPA {probation_tier.condition} {probation_tier.bound}",
            )

        # Study Hour Reduction Key Information
        start_row = 21
        start_column = 10
        for i, reduce_hours_tier in enumerate(
            (
                reduce_one_tier,
                reduce_two_tiers,
            )
        ):
            worksheet.cell(
                row=start_row + i,
                column=start_column,
                value=reduce_hours_tier.description_in_house,
            )
            worksheet.cell(
                row=start_row + i,
                column=start_column + 2,
                value="GPA "
                + f"{reduce_hours_tier.term_condition.condition} "
                + f"{reduce_hours_tier.term_condition.bound}",
            )
            worksheet.cell(
                row=start_row + i,
                column=start_column + 4,
                value="GPA "
                + f"{reduce_hours_tier.cumulative_condition.condition} "
                + f"{reduce_hours_tier.cumulative_condition.bound}",
            )

        # Social Probation and No Punting Reduction Key Information
        start_row = 23
        start_column = 10
        for i, reduce_probation_tier in enumerate(
            (
                get_off_no_punting,
                get_off_social_probation,
            )
        ):
            worksheet.cell(
                row=start_row + i,
                column=start_column,
                value=reduce_probation_tier.description_in_house,
            )
            worksheet.cell(
                row=start_row + i,
                column=start_column + 2,
                value=f"GPA {reduce_probation_tier.condition} "
                + f"{reduce_probation_tier.bound}",
            )

    def _organize_report_data(self, selected_term: str) -> None:
        """Organizes the report information.

        Args:
            selected_term (str): The desired term to organize data from.

        Returns:
            None: None
        """

        def sort_key(
            member_data: GradeReportItem,
        ) -> tuple[bool, float | int, float | int, str]:
            """Function used for sorting grade report row values.

            Rows are sorted by: out-of-house before in-house, pledge class
            year, pledge class semester (Spring before Fall), member last name.

            Gives the row value from a list a number value used for sorting.

            Args:
                member_data (GradeReportItem): The dataclass that represents a
                    row.

            Returns:
                tuple[bool, float | int, float | int, str]: The sorting value
                    for the row.
            """
            pledge_class_year_order: float | int = float("inf")
            pledge_class_semester_order: float | int = float("inf")
            if member_data.pledge_class is not None:
                pledge_class_year_order = int(member_data.pledge_class[2:])
                if member_data.pledge_class[:2] == "SP":
                    pledge_class_semester_order = 0
                else:
                    pledge_class_semester_order = 1

            return (
                not member_data.out_of_house,
                pledge_class_year_order,
                pledge_class_semester_order,
                member_data.name.split()[1],
            )

        report_data: list[GradeReportItem] = []
        for member in self.members.values():
            member_data: GradeReportItem = member.determine_academic_probation(
                selected_term
            )
            report_data.append(member_data)
        self.report_data = sorted(report_data, key=sort_key)

    def _write_member_data(self, selected_term: str) -> None:
        """Writes the report data.

        Args:
            selected_term (str): The desired term to write study hour data
                from.

        Raises:
            ValueError: When the method is called when the workbook attribute
                is None.

        Returns:
            None: None
        """
        self._organize_report_data(selected_term)

        if self.workbook is None:
            raise ValueError("Workbook cannot be None, copy template first.")
        worksheet = self.workbook.active

        start_row = 2
        start_column = 1
        for i, row in enumerate(self.report_data):
            if row.cumulative_gpa is not None:
                row.cumulative_gpa = float(f"{row.cumulative_gpa:.3f}")
            if row.previous_gpa is not None:
                row.previous_gpa = float(f"{row.previous_gpa:.3f}")
            if row.term_gpa is not None:
                row.term_gpa = float(f"{row.term_gpa:.3f}")
            for j, value in enumerate(
                (
                    row.name,
                    row.pledge_class,
                    row.cumulative_gpa,
                    row.previous_term,
                    row.previous_gpa,
                    row.term,
                    row.term_gpa,
                    row.study_hours,
                )
            ):
                worksheet.cell(
                    row=start_row + i, column=start_column + j, value=value
                )

    def _write_statistics(self, selected_term: str) -> None:
        """Writes statistics about the grade data to the report.

        Args:
            selected_term (str): _description_

        Raises:
            ValueError: When the method is called when the workbook attribute
                is None.

        Returns:
            None: None
        """
        if not self.report_data:
            self._organize_report_data(selected_term)

        if self.workbook is None:
            raise ValueError("Workbook cannot be None, copy template first.")
        worksheet = self.workbook.active

        gpa_sum: float = 0.00
        num_members: int = 0

        # Gather report data by pledge class.
        pledge_class_gpa_statistics: dict[str, list[GradeReportItem]] = dict()
        for row in self.report_data:
            if row.term_gpa is not None:
                gpa_sum += row.term_gpa
                num_members += 1
            if row.pledge_class is None:
                continue
            pledge_class_semester = Semester(row.pledge_class)
            if pledge_class_semester.type == "SP":
                pledge_class_semester.type = "FS"
                pledge_class_semester.year -= 1
            if repr(pledge_class_semester) not in pledge_class_gpa_statistics:
                pledge_class_gpa_statistics[repr(pledge_class_semester)] = []
            pledge_class_gpa_statistics[repr(pledge_class_semester)].append(
                row
            )

        start_row = 26
        start_col = 14
        house_average_gpa: float | None = (
            gpa_sum / num_members if num_members > 0 else None
        )
        worksheet.cell(
            row=start_row,
            column=start_col,
            value=(
                f"{house_average_gpa:.3f}"
                if house_average_gpa is not None
                else None
            ),
        )

        # Write GPA statistics to Sheet.
        start_row = 30
        start_column = 10
        pledge_class_average_term_gpas: dict[str, float | None] = dict()
        for i, pledge_class in enumerate(pledge_class_gpa_statistics):
            # Collect Term GPAs.
            term_gpas: list[float] = []
            for row in pledge_class_gpa_statistics[pledge_class]:
                if row.term_gpa is None:
                    continue
                term_gpas.append(row.term_gpa)
            # Get Average Term GPA for the class.
            average_term_gpa: float | None = None
            if len(term_gpas) > 0:
                average_term_gpa = sum(term_gpas) / len(term_gpas)
                average_term_gpa = float(f"{average_term_gpa:.3f}")

            pledge_class_average_term_gpas[pledge_class] = average_term_gpa

            # Collet Previous Term GPAs.
            previous_term_gpas: list[float] = []
            for row in pledge_class_gpa_statistics[pledge_class]:
                if row.previous_gpa is None:
                    continue
                previous_term_gpas.append(row.previous_gpa)
            # Get Average Previous Term GPA for the class.
            average_previous_term_gpa: float | None = None
            if len(previous_term_gpas) > 0:
                average_previous_term_gpa = sum(previous_term_gpas) / len(
                    previous_term_gpas
                )
                average_previous_term_gpa = float(
                    f"{average_previous_term_gpa:.3f}"
                )

            # Calculate the Change in GPA from previous to current term.
            change_in_gpa: float | None = None
            if (
                average_term_gpa is not None
                and average_previous_term_gpa is not None
            ):
                change_in_gpa = average_term_gpa - average_previous_term_gpa
                change_in_gpa = float(f"{change_in_gpa:.3f}")

            # Write Data to Sheet.
            for j, value in enumerate(
                (
                    pledge_class,
                    len(pledge_class_gpa_statistics[pledge_class]),
                    average_term_gpa,
                    average_previous_term_gpa,
                    change_in_gpa,
                )
            ):
                worksheet.cell(
                    row=start_row + i,
                    column=start_column + j,
                    value=value,
                )

        # Write Pledge Class GPA ranking to sheet
        def rank_key(pledge_class: str) -> float:
            """Generates a rank value for the pledge class GPA data.

            Args:
                pledge_class (str): The pledge class whose average GPA is
                    being ranked.

            Returns:
                float: The sorting rank value.
            """
            rank = pledge_class_average_term_gpas.get(pledge_class)
            if rank is None:
                return -1
            return rank

        pledge_class_rankings: dict[str, int] = dict()
        for rank, pledge_class in enumerate(
            sorted(pledge_class_average_term_gpas, key=rank_key, reverse=True)
        ):
            pledge_class_rankings[pledge_class] = rank + 1

        start_row = 30
        start_column = 15
        for i, pledge_class in enumerate(pledge_class_average_term_gpas):
            if pledge_class_average_term_gpas[pledge_class] is None:
                continue
            worksheet.cell(
                row=start_row + i,
                column=start_column,
                value=f"{pledge_class_rankings[pledge_class]}",
            )

    def _save(self, save_path: str, open_file: bool = False) -> None:
        """Saves the report to file.

        Args:
            save_path (str): The file path where the report is saved to.
            open_file (bool, optional): Option to open the file after it is
                saved. Defaults to False.

        Raises:
            ValueError: When the method is called when the workbook attribute
                is None.

        Returns:
            None: None
        """
        if self.workbook is None:
            raise ValueError("Workbook cannot be None, copy template first.")
        self.workbook.save(save_path)
        if open_file:
            os.startfile(save_path)
