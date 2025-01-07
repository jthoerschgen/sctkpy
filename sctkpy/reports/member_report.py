"""Object for generating Member Grade Reports"""

import os
import shutil
from datetime import datetime

import openpyxl

from sctkpy.config import (
    member_report_template_path,
    proj_root_dir,
    strike,
    super_strike,
)
from sctkpy.member.course import Course
from sctkpy.member.member import Member
from sctkpy.member.term import Term
from sctkpy.reports.grade_report import GradeReport


class MemberReport(GradeReport):
    """Extra logic for writing a Member Grade Report from a GradeReport object.

    Attributes:
        template_path (str): Path of the template spreadsheet.
        save_dir (str): Directory path of where the report is saved.
        workbook (openpyxl.Workbook): The workbook where the report is written
            to. Defaults to None.
    """

    def __init__(self, roster_report_path: str, save_dir: str = proj_root_dir):
        super().__init__(roster_report_path)
        self.template_path: str = member_report_template_path
        self.save_dir: str = save_dir
        self.workbook: openpyxl.Workbook | None = None

    def __call__(self, open_on_finish: bool = True) -> None:
        current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        save_path: str = os.path.join(
            self.save_dir, f"Member_Report_{current_time}.xlsx"
        )

        self._copy_template(save_path)
        self._write_member_data()
        self._save(save_path, open_on_finish)

    def _copy_template(self, save_path: str) -> None:
        """Makes a copy of the member report template.

        Args:
            save_path (str): The path where the copy is saved.

        Returns:
            None: None
        """
        shutil.copy(src=self.template_path, dst=save_path)
        self.workbook = openpyxl.load_workbook(save_path)

    def _write_member_data(self) -> None:
        """Writes the report data.

        Raises:
            ValueError: When the method is called when the workbook attribute
                is None.
            ValueError: If there are not the expected template sheetnames in
                the workbook.

        Returns:
            None: None
        """

        def members_sort_key(
            member: Member,
        ) -> tuple[float | int, float | int, str]:
            """Function used for sorting Members.

            Items are sorted by: pledge class year, pledge class semester
                (Spring before Fall), member last name.

            Gives the member value from a list a number value used for sorting.

            Args:
                member (Member): The member item being sorted.

            Returns:
                tuple[float | int, float | int, str]: The sorting value
                    for the member.
            """
            pledge_class = member.get_pledge_class()
            pledge_class_year_order: float | int = float("inf")
            pledge_class_semester_order: float | int = float("inf")
            if pledge_class is not None:
                pledge_class_year_order = int(pledge_class[2:])
                if pledge_class[:2] == "SP":
                    pledge_class_semester_order = 0
                else:
                    pledge_class_semester_order = 1
            return (
                pledge_class_year_order,
                pledge_class_semester_order,
                member.name,
            )

        def term_sort_key(term: Term) -> tuple[int, int]:
            """Function used for sorting Terms.

            Items are sorted by: year and semester, (SP before FS).

            Gives the term value from a list a number value used for sorting.

            Args:
                term (Term): The term item that is being sorted.

            Returns:
                tuple[int, int]: The sorting value for the term.
            """
            term_year: int = term.term.year
            term_semester_order: int = 0 if term.term.type == "SP" else 1
            return term_year, term_semester_order

        def course_sort_key(course: Course) -> tuple[str, str]:
            """Function used for sorting Courses.

            Items are sorted by: catalog number and course name.

            Gives the course value from a list a number value used for sorting.

            Args:
                course (Course): The course item that is being sorted.

            Returns:
                tuple[str, str]: The sorting value for the course.
            """
            return course.catalog_number, course.name

        if self.workbook is None:
            raise ValueError("Workbook cannot be None, copy template first.")
        for sheetname in ("Overview Info", "Individual Info Template"):
            if sheetname not in self.workbook.sheetnames:
                raise ValueError(
                    f"Invalid template. {sheetname} not a worksheet in "
                    + f"template: {self.template_path}."
                )

        overview_worksheet = self.workbook["Overview Info"]

        # Write Key Information
        key_information_start_row: int = 3
        key_information_start_column: int = 2
        for i, strike_type in enumerate((strike, super_strike)):
            for j, strike_value in enumerate(
                (
                    strike_type.description_in_house,
                    strike_type.number_chances,
                    f"GPA {strike_type.condition} {strike_type.bound}",
                )
            ):
                overview_worksheet.cell(
                    row=key_information_start_row + i,
                    column=key_information_start_column + j,
                    value=strike_value,
                )

        overview_start_row: int = 6

        for overview_row_offset, member in enumerate(
            sorted(self.members.values(), key=members_sort_key)
        ):
            member_worksheet = self.workbook.copy_worksheet(
                self.workbook["Individual Info Template"]
            )
            member_worksheet.title = member.name

            # Starting Rows and Columns
            courses_start_row: int = 5
            courses_start_column: int = 1

            terms_start_row: int = 5
            terms_start_column: int = 6

            # Create Row for Member in the Overview Worksheet
            overview_worksheet.cell(
                row=overview_start_row + overview_row_offset,
                column=1,
                value=member.name,
            )
            overview_worksheet.cell(
                row=overview_start_row + overview_row_offset,
                column=2,
                value=member.get_pledge_class(),
            )
            overview_worksheet.cell(
                row=overview_start_row + overview_row_offset,
                column=3,
                value=(
                    "=COUNTA(INDIRECT("
                    + f'"\'" & A{overview_start_row + overview_row_offset} '
                    + '& "\'!J5:J100"))'
                ),
            )  # Add formula for counting terms.
            overview_worksheet.cell(
                row=overview_start_row + overview_row_offset,
                column=4,
                value=(
                    "=COUNTA(INDIRECT("
                    + f'"\'" & A{overview_start_row + overview_row_offset} '
                    + '& "\'!K5:K100"))'
                ),
            )  # Add formula for counting strikes.
            overview_worksheet.cell(
                row=overview_start_row + overview_row_offset,
                column=5,
                value=(
                    "=COUNTA(INDIRECT("
                    + f'"\'" & A{overview_start_row + overview_row_offset} '
                    + '& "\'!F5:F100"))'
                ),
            )  # Add formula for counting super strikes.

            # Write Member Name to their own Worksheet
            member_worksheet.cell(
                row=2, column=2, value=member.name.split()[0]
            )
            member_worksheet.cell(
                row=2, column=3, value=" ".join(member.name.split()[1:])
            )

            for term_row_offset, term in enumerate(
                sorted(member.terms.values(), key=term_sort_key)
            ):
                # Calculate Term Credit Hour Load
                term_credit_hour_load: float = 0.0
                for course in term.courses:
                    term_credit_hour_load += course.hours

                # Calculate Term GPA
                term_gpa: float | None = term.calculate_gpa()
                if term_gpa is not None:
                    term_gpa = float(f"{term_gpa:.3f}")

                if term.term_cum_gpa is not None:
                    term.term_cum_gpa = float(f"{term.term_cum_gpa:.3f}")

                # Determine if Member has Strike for the Term
                has_strike: bool = False
                has_super_strike: bool = False
                if term_gpa is not None:
                    has_strike = strike.test_tier(term_gpa)
                    has_super_strike = super_strike.test_tier(term_gpa)

                # Write Term Information to Worksheet
                for term_column_offset, term_value in enumerate(
                    (
                        str(term.term),
                        term_gpa,
                        term.term_cum_gpa,
                        term_credit_hour_load,
                        "X" if has_strike else "",
                        "X" if has_super_strike else "",
                    )
                ):
                    member_worksheet.cell(
                        row=terms_start_row + term_row_offset,
                        column=terms_start_column + term_column_offset,
                        value=term_value,
                    )

                for course_row_offset, course in enumerate(
                    sorted(term.courses, key=course_sort_key)
                ):
                    # Write Course Information to Worksheet
                    for course_column_offset, course_value in enumerate(
                        (
                            str(term.term),
                            course.catalog_number,
                            course.name,
                            course.hours,
                            (
                                course.grade.name
                                if course.grade is not None
                                else ""
                            ),
                        )
                    ):
                        member_worksheet.cell(
                            row=courses_start_row + course_row_offset,
                            column=courses_start_column + course_column_offset,
                            value=course_value,
                        )
                courses_start_row += len(term.courses)
        self.workbook.remove(
            self.workbook["Individual Info Template"]
        )  # Remove template worksheet when done.
        self.workbook.active = overview_worksheet

    def _save(self, save_path: str, open_file: bool = False) -> None:
        """Saves the report to file.

        Args:
            save_path (str): The file path where the report is saved to.
            open_file (bool, optional): Option to open the file after it is
                saved. Defaults to False.

        Raises:
            ValueError: When the method is called when the workbook attribute
                is None.

        Returns:
            None: None
        """
        if self.workbook is None:
            raise ValueError("Workbook cannot be None, copy template first.")
        self.workbook.save(save_path)
        if open_file:
            os.startfile(save_path)
