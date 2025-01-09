"""Object for generating Files Due Reports"""

import os
import shutil
from datetime import datetime

import openpyxl

from sctkpy.config import files_hit_list_template_path
from sctkpy.reports.grade_report import GradeReport


class FilesReport(GradeReport):
    """Extra logic for writing a Files Due Report from a GradeReport object.

    Attributes:
        template_path (str): Path of the template spreadsheet.
        save_dir (str): Directory path of where the report is saved.
        report_data (dict[tuple[str, str], set[str]]): Dictionary where the
            key is the course number and name and the value is a set of names
            of the member's in the course. Defaults to an empty dictionary.
        workbook (openpyxl.Workbook): The workbook where the report is written
            to. Defaults to None.
    """

    def __init__(self, roster_report_path: str, save_dir: str):
        super().__init__(roster_report_path)
        self.template_path = files_hit_list_template_path
        self.save_dir = save_dir
        self.report_data: dict[tuple[str, str], set[str]] = dict()
        self.workbook: openpyxl.Workbook | None = None

    def __call__(
        self,
        selected_term: str,
        open_on_finish: bool = True,
    ) -> None:
        current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        save_path: str = os.path.join(
            self.save_dir, f"Files_Report_{selected_term}_{current_time}.xlsx"
        )

        self._copy_template(save_path)
        self._write_files_data(selected_term)
        self._save(save_path, open_on_finish)

    def _copy_template(self, save_path: str) -> None:
        """Makes a copy of the files due list report template.

        Args:
            save_path (str): The path where the copy is saved.

        Returns:
            None: None
        """
        shutil.copy(src=self.template_path, dst=save_path)
        self.workbook = openpyxl.load_workbook(save_path)

    def _organize_report_data(self, selected_term: str) -> None:
        """Organizes the report information.

        Args:
            selected_term (str): The desired term to organize data from.

        Returns:
            None: None
        """
        report_data: dict[tuple[str, str], set[str]] = dict()
        for member in self.members.values():
            for course in member.terms[selected_term].courses:
                course_id: tuple[str, str] = (
                    course.catalog_number,
                    course.name,
                )
                if course_id not in report_data:
                    report_data[course_id] = set()
                report_data[course_id].add(member.name)
        self.report_data = report_data

    def _write_files_data(self, selected_term: str) -> None:
        """Writes the report data.

        Args:
            selected_term (str): The desired term to write study hour data
                from.

        Raises:
            ValueError: When the method is called when the workbook attribute
                is None.

        Returns:
            None: None
        """
        self._organize_report_data(selected_term)

        if self.workbook is None:
            raise ValueError("Workbook cannot be None, copy template first.")
        worksheet = self.workbook.active

        # Write Title information
        start_row = 1
        start_col = 1
        worksheet.cell(
            row=start_row,
            column=start_col,
            value=f"File Responsibility Report for: {selected_term}",
        )
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        worksheet.cell(
            row=start_row + 1,
            column=start_col,
            value=f"Report Generated at {current_time}",
        )
        worksheet.cell(
            row=start_row + 2,
            column=start_col,
            value=f"Total: {len(self.report_data)} files",
        )

        # Write Files Due List Data
        start_row = 6
        start_col = 1
        i = 0
        for (course_number, course_name), members in self.report_data.items():
            worksheet.cell(
                row=start_row + i,
                column=start_col,
                value=course_name,
            )
            worksheet.cell(
                row=start_row + i,
                column=start_col + 1,
                value=course_number,
            )
            i += 1
            for dude in members:
                worksheet.cell(
                    row=start_row + i,
                    column=start_col + 4,
                    value=dude,
                )
                i += 1

    def _save(self, save_path: str, open_file: bool = False) -> None:
        """Saves the report to file.

        Args:
            save_path (str): The file path where the report is saved to.
            open_file (bool, optional): Option to open the file after it is
                saved. Defaults to False.

        Raises:
            ValueError: When the method is called when the workbook attribute
                is None.

        Returns:
            None: None
        """
        if self.workbook is None:
            raise ValueError("Workbook cannot be None, copy template first.")
        self.workbook.save(save_path)
        if open_file:
            os.startfile(save_path)
