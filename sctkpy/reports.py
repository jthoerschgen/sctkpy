# -*- coding: utf-8 -*-

"""Generate reports using templates the sheets in ./templates
"""

import datetime
import os
import shutil

import openpyxl
import openpyxl.drawing
import openpyxl.drawing.image
from PIL import Image
from tqdm import tqdm

from .config import (  # cumulative_gpa_strike,
    files_hit_list_template_path,
    grade_report_template_path,
    individual_report_template_path,
    no_punting,
    no_study_hours,
    proj_root_dir,
    social_probation,
    study_checks_template_path,
    term_gpa_strike,
    term_gpa_super_strike,
    tier_one,
    tier_two,
)
from .core import get_member_info, plot_member_gpa
from .db_funcs import get_files_due_list_db
from .member import Member


def generate_grade_report(term: str, open_on_finish: bool = True) -> None:
    """Make a copy of and write data to a grade report for grade data from a
    given term

    Args:
        term (str): Term to use grade data from
    """
    current_time = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    grade_report_save_path: str = os.path.join(
        proj_root_dir, f"Grade_Report_{term}_{current_time}.xlsx"
    )
    shutil.copy(src=grade_report_template_path, dst=grade_report_save_path)

    workbook = openpyxl.load_workbook(filename=grade_report_save_path)
    worksheet = workbook.active

    # Write member grade data
    start_row = 2
    start_col = 1
    members: list[Member] = get_member_info()

    statistics_data: dict = {"pledge_class": [], "gpa": []}

    for i, member in enumerate(members):  # Write columns
        ret_term, term_gpa, study_hours = member.study_hours(term)

        if term_gpa is not None:
            statistics_data["pledge_class"].append(member.pledge_class)
            statistics_data["gpa"].append(term_gpa)

        rounded_term_gpa = (
            float(f"{term_gpa:.3f}") if term_gpa is not None else None
        )

        for j, datum in enumerate(
            (
                member.name,
                ret_term,
                member.pledge_class,
                rounded_term_gpa,
                study_hours,
            )
        ):  # Write row
            worksheet.cell(
                row=start_row + i, column=start_col + j, value=datum
            )

    # House GPA Statistics

    # Whole house average GPA
    house_avg_gpa: float = sum(statistics_data["gpa"]) / len(
        statistics_data["gpa"]
    )
    start_row = 17
    start_col = 9
    worksheet.cell(
        row=start_row, column=start_col, value=f"{house_avg_gpa:.3f}"
    )

    unique_pledge_classes = []
    for pledge_class in statistics_data["pledge_class"]:
        if pledge_class not in unique_pledge_classes:
            unique_pledge_classes.append(pledge_class)

    # Pledge class average GPA
    start_row = 21
    start_col = 7
    for i, pledge_class in enumerate(unique_pledge_classes):
        pc_grade_data = [
            statistics_data["gpa"][i]
            for i, pc in enumerate(statistics_data["pledge_class"])
            if pc == pledge_class
        ]
        pc_avg_grade = sum(pc_grade_data) / len(pc_grade_data)

        # Write data
        worksheet.cell(row=start_row + i, column=start_col, value=pledge_class)
        worksheet.cell(
            row=start_row + i, column=start_col + 1, value=len(pc_grade_data)
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 2,
            value=f"{pc_avg_grade:.3f}",
        )

    # Add hidden numbers used for comparison in the xlsx conditional formatting
    # In column F, yellow => tier one (F4), red => tier two (F5).
    start_row = 4
    start_col = 6
    for i, option in enumerate((tier_one, tier_two)):
        worksheet.cell(
            row=start_row + i,
            column=start_col,
            value=option.bound,
        )

    # In-House - Key Info
    start_row = 4
    start_col = 7
    for i, option in enumerate(
        (no_study_hours, tier_one, tier_two, no_punting)
    ):
        worksheet.cell(
            row=start_row + i,
            column=start_col,
            value=option.desc_in_house,
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 1,
            value=option.result_in_house,
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 2,
            value=f"GPA {option.condition} {option.bound}",
        )

    # Out-House - Key Info
    start_row = 10
    start_col = 7
    for i, option in enumerate((no_study_hours, tier_one, tier_two)):
        worksheet.cell(
            row=start_row + i,
            column=start_col,
            value=option.desc_out_house,
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 1,
            value=option.result_out_house,
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 2,
            value=f"GPA {option.condition} {option.bound}",
        )

    # Social Probation - Key Info
    start_row = 15
    start_col = 7
    option = social_probation
    worksheet.cell(
        row=start_row,
        column=start_col,
        value=option.desc_in_house,
    )
    worksheet.cell(
        row=start_row,
        column=start_col + 1,
        value=option.result_in_house,
    )
    worksheet.cell(
        row=start_row,
        column=start_col + 2,
        value=f"GPA {option.condition} {option.bound}",
    )

    workbook.save(grade_report_save_path)
    if open_on_finish:
        os.startfile(grade_report_save_path)

    return


def generate_study_check_sheet(term: str, open_on_finish: bool = True) -> None:
    """Make a copy of and write data to a study-checklist for grade data from
    a given term

    Args:
        term (str): Term to use grade data from
    """
    current_time = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    study_checks_save_path: str = os.path.join(
        proj_root_dir, f"Study_Check_Sheet_{term}_{current_time}.xlsx"
    )
    shutil.copy(src=study_checks_template_path, dst=study_checks_save_path)

    workbook = openpyxl.load_workbook(filename=study_checks_save_path)
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value=f"{term} Week of:")

    start_row = 3
    start_col = 1
    members: list[Member] = get_member_info()

    i = 0
    for member in members:  # Write columns
        _, _, study_hours = member.study_hours(term)
        if len(study_hours) > 0 and not member.out_of_house:
            worksheet.cell(
                row=start_row + i, column=start_col, value=member.name
            )
            worksheet.cell(
                row=start_row + i, column=start_col + 1, value=study_hours
            )
            i += 1

    workbook.save(study_checks_save_path)
    if open_on_finish:
        os.startfile(study_checks_save_path)

    return


def generate_files_responsibility_report(
    term: str, open_on_finish: bool = True
) -> None:
    """Make a copy of and write data to a files turn-in checklist from course
    data from a given term

    Args:
        term (str): Term to use course data from
    """
    current_time = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    files_hit_list_save_path: str = os.path.join(
        proj_root_dir, f"Files_Hit_List_{term}_{current_time}.xlsx"
    )
    shutil.copy(src=files_hit_list_template_path, dst=files_hit_list_save_path)

    workbook = openpyxl.load_workbook(filename=files_hit_list_save_path)
    worksheet = workbook.active

    files_due_list = get_files_due_list_db(term)

    # Title information
    start_row = 1
    start_col = 1
    worksheet.cell(
        row=start_row,
        column=start_col,
        value=f"File Responsibility Report for: {term}",
    )
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    worksheet.cell(
        row=start_row + 1,
        column=start_col,
        value=f"Report Generated at {current_time}",
    )
    worksheet.cell(
        row=start_row + 2,
        column=start_col,
        value=f"Total: {len(files_due_list)} files",
    )

    # Files Due List Data
    start_row = 6
    start_col = 1
    i = 0
    for class_name, class_num, dudes in files_due_list:
        worksheet.cell(
            row=start_row + i,
            column=start_col,
            value=class_name,
        )
        worksheet.cell(
            row=start_row + i,
            column=start_col + 1,
            value=class_num,
        )
        i += 1
        for dude in dudes.split(","):
            worksheet.cell(
                row=start_row + i,
                column=start_col + 4,
                value=dude,
            )
            i += 1

    workbook.save(files_hit_list_save_path)
    if open_on_finish:
        os.startfile(files_hit_list_save_path)

    return


def generate_individual_academic_report(open_on_finish: bool = True) -> None:
    """Generate an extensive report on each member's academics over all data

    Args:
        open_on_finish (bool): Open file automatically when done
    """
    current_time = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    individual_report_save_path: str = os.path.join(
        proj_root_dir, f"Individual_Grade_Report_{current_time}.xlsx"
    )
    shutil.copy(
        src=individual_report_template_path, dst=individual_report_save_path
    )

    workbook = openpyxl.load_workbook(filename=individual_report_save_path)

    start_row = 2
    start_col = 5
    # Key information
    for i, strike_type in enumerate((term_gpa_strike, term_gpa_super_strike)):
        workbook["info"].cell(
            row=start_row + i,
            column=start_col,
            value=strike_type.desc_in_house + ":",
        )
        workbook["info"].cell(
            row=start_row + i,
            column=start_col + 2,
            value=strike_type.condition + str(strike_type.bound),
        )
        workbook["info"].cell(
            row=start_row + i,
            column=start_col + 3,
            value=strike_type.num_chances,
        )

    info_sheet_start_row = 5
    info_sheet_start_col = 1
    members: list[Member] = get_member_info()
    for member_index, member in tqdm(
        list(enumerate(members)), desc="Generating Individual GPA Reports"
    ):
        worksheet = workbook.copy_worksheet(workbook["template"])
        worksheet.title = member.name

        for cf_range in workbook[
            "template"
        ].conditional_formatting._cf_rules:  # pylint: disable=protected-access
            for cf_rule in workbook[
                "template"
            ].conditional_formatting._cf_rules[
                cf_range
            ]:  # pylint: disable=protected-access
                worksheet.conditional_formatting.add(cf_range, cf_rule)

        start_row = 1
        start_col = 4
        for i, name_part in enumerate(
            (
                member.name.split()[0],
                (
                    " ".join(member.name.split()[1:])
                    if len(member.name.split()) > 1
                    else ""
                ),
            ),
        ):
            workbook["info"].cell(
                row=info_sheet_start_row + member_index,
                column=info_sheet_start_col + i,
                value=name_part,
            )  # Write to main info sheet
            worksheet.cell(
                row=start_row,
                column=start_col + i,
                value=name_part,
            )  # Write to their individual report

        terms_list, term_gpa_list, cumulative_gpa_list = (
            member.generate_academic_report()
        )

        # Write to info sheet
        workbook["info"].cell(
            row=info_sheet_start_row + member_index,
            column=info_sheet_start_col + 2,
            value=member.pledge_class,
        )  # Write to main info sheet
        workbook["info"].cell(
            row=info_sheet_start_row + member_index,
            column=info_sheet_start_col + 3,
            value=len(
                [
                    gpa
                    for gpa in term_gpa_list
                    if gpa is not None and term_gpa_strike.test_tier(gpa)
                ]
            ),
        )  # Number of strikes
        workbook["info"].cell(
            row=info_sheet_start_row + member_index,
            column=info_sheet_start_col + 4,
            value=len(
                [
                    gpa
                    for gpa in term_gpa_list
                    if gpa is not None and term_gpa_super_strike.test_tier(gpa)
                ]
            ),
        )  # Number of super strikes
        workbook["info"].cell(
            row=info_sheet_start_row + member_index,
            column=info_sheet_start_col + 5,
            value=len(terms_list),
        )  # Number of terms

        # Write Summary Information
        start_row = 4
        start_col = 6
        for i, term_gpa_data in enumerate(
            zip(terms_list, term_gpa_list, cumulative_gpa_list)
        ):
            term, term_gpa, cumulative_gpa = term_gpa_data

            rounded_term_gpa = (
                float(f"{term_gpa:.3f}") if term_gpa is not None else None
            )
            rounded_cumulative_gpa = (
                float(f"{cumulative_gpa:.3f}")
                if cumulative_gpa is not None
                else None
            )

            worksheet.cell(
                row=start_row + i,
                column=start_col,
                value=term,
            )  # Term name
            worksheet.cell(
                row=start_row + i,
                column=start_col + 1,
                value=rounded_term_gpa,
            )  # Calculated term GPA
            campus_term_gpa = (
                member.terms[term].term_gpa
                if isinstance(member.terms[term].term_gpa, float)
                else None
            )
            worksheet.cell(
                row=start_row + i,
                column=start_col + 2,
                value=(
                    float(f"{campus_term_gpa:.3f}")
                    if campus_term_gpa is not None
                    else ""
                ),
            )  # Campus term GPA
            worksheet.cell(
                row=start_row + i,
                column=start_col + 3,
                value=rounded_cumulative_gpa,
            )  # Calculated cumulative GPA
            worksheet.cell(
                row=start_row + i,
                column=start_col + 4,
                value=float(f"{member.terms[term].term_cum_gpa:.3f}"),
            )  # Campus cumulative GPA
            worksheet.cell(
                row=start_row + i,
                column=start_col + 5,
                value=sum(
                    [course.hrs for course in member.terms[term].classes]
                ),
            )  # Credit hours for the semester
            worksheet.cell(
                row=start_row + i,
                column=start_col + 6,
                value=(
                    "X"
                    if (
                        #     cumulative_gpa is not None
                        #     and cumulative_gpa_strike.test_tier(cumulative_gpa)
                        # )
                        # or (
                        term_gpa is not None
                        and term_gpa_strike.test_tier(term_gpa)
                    )
                    else None
                ),
            )  # Regular strike
            worksheet.cell(
                row=start_row + i,
                column=start_col + 7,
                value=(
                    "X"
                    if (
                        term_gpa is not None
                        and term_gpa_super_strike.test_tier(term_gpa)
                    )
                    else None
                ),
            )  # Super strike

        # Write course information
        start_row = 4
        for term in terms_list:
            for course in member.terms[term].classes:
                start_col = 1
                for i, col_val in enumerate(
                    (
                        term,
                        course.class_name,
                        course.catalog_no,
                        course.hrs,
                        course.grade,
                    )
                ):
                    worksheet.cell(
                        row=start_row,
                        column=start_col + i,
                        value=col_val,
                    )
                start_row += 1

        # Add matplotlib plot PNG image of GPA over time to sheet
        plot_image: Image.Image | None = plot_member_gpa(member)
        if plot_image is not None:
            plot = openpyxl.drawing.image.Image(plot_image)
            plot.anchor = "N4"
            worksheet.add_image(plot)

    workbook.remove(workbook["template"])  # remove blank template sheet

    workbook.save(individual_report_save_path)
    if open_on_finish:
        os.startfile(individual_report_save_path)

    return
